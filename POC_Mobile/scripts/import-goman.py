import json
import re
from pathlib import Path

import openpyxl

path = r"c:\Users\Italo\OneDrive - GRUPO CGB\Área de Trabalho\BANCO - GOMAN.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["GOMAN"]


def slug(text: str) -> str:
    t = text.lower()
    t = re.sub(r"[^a-z0-9]+", "-", t)
    return t.strip("-")


items = []
cat_order: list[str] = []
cat_map: dict[str, dict] = {}
idx = 1

for r in range(2, ws.max_row + 1):
    cat = ws.cell(r, 1).value
    pergunta = ws.cell(r, 3).value
    peso = ws.cell(r, 4).value
    grav = ws.cell(r, 5).value
    if not pergunta:
        continue

    cat = str(cat).strip()
    pergunta = str(pergunta).replace("\n", " ").strip()
    grav = str(grav).strip() if grav else ""

    if cat not in cat_map:
        cat_map[cat] = {"id": slug(cat), "label": cat, "perguntas": []}
        cat_order.append(cat)

    cat_map[cat]["perguntas"].append(
        {
            "id": f"goman-{idx:03d}",
            "texto": pergunta,
            "peso": int(peso) if peso else 0,
            "gravidade": grav,
        }
    )
    idx += 1

categorias = [cat_map[c] for c in cat_order]

out = Path(__file__).resolve().parent.parent / "src" / "data" / "goman-checklist.ts"
header = """export type Gravidade = "Leve" | "Médio" | "Grave"
export type RespostaChecklist = "conforme" | "nao_conforme" | null

export interface PerguntaGoman {
  id: string
  texto: string
  peso: number
  gravidade: Gravidade
}

export interface CategoriaGoman {
  id: string
  label: string
  perguntas: PerguntaGoman[]
}

export interface RespostaPergunta {
  perguntaId: string
  resposta: Exclude<RespostaChecklist, null>
  observacao?: string
}

export const gomanChecklist: CategoriaGoman[] = """

footer = """
export const totalPerguntasGoman = gomanChecklist.reduce(
  (n, c) => n + c.perguntas.length,
  0,
)
"""

content = header + json.dumps(categorias, ensure_ascii=False, indent=2) + footer
out.write_text(content, encoding="utf-8")

print(f"Written {out}")
print(f"Categories: {len(categorias)}, Questions: {idx - 1}")
for c in categorias:
    print(f"  - {c['label']}: {len(c['perguntas'])}")
